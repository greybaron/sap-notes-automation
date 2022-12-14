import openpyxl


def analysis(system, india_xlsx_path, notes_from_sap): 
       
    notes_in_xlsx = set()

    ### exc handling in GUI
    workbook_india = openpyxl.load_workbook(india_xlsx_path) #wb from india BUT FORMATTED
    worksheet_india = None
   
    # finding the exact name
    for sheet in workbook_india.sheetnames:
        if worksheet_india is not None:
            break

        if system in sheet:
            worksheet_india = workbook_india[sheet]

    if worksheet_india is None:
        raise IndexError(f"\n\nSheet '{system}' could not be found in workbook '{india_xlsx_path}'")


    # get notes from India XSLX, save to set 'notes_in_xlsx'
    for e in range (2, worksheet_india.max_row + 1):
        cell_address = 'A' + str(e)
        cell_value = worksheet_india[cell_address].value
        if cell_value is not None:
            notes_in_xlsx.add(cell_value)

    only_in_sap = notes_from_sap - notes_in_xlsx
    only_in_xlsx = notes_in_xlsx - notes_from_sap


    return only_in_sap, only_in_xlsx
